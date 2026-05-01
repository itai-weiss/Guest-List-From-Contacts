from __future__ import annotations

from io import BytesIO
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook

from flask_app import SESSIONS, create_app


@pytest.fixture(autouse=True)
def clear_sessions() -> None:
    SESSIONS.clear()
    yield
    SESSIONS.clear()


def _make_workbook_bytes() -> bytes:
    output = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Guests"
    worksheet.append(["שם"])
    worksheet.append(["Ori Bina"])
    workbook.save(output)
    output.seek(0)
    return output.read()


def _make_contacts_vcf() -> bytes:
    return (
        "BEGIN:VCARD\n"
        "VERSION:3.0\n"
        "FN:Ori Bina\n"
        "N:Bina;Ori;;;\n"
        "TEL;TYPE=CELL:0525665352\n"
        "END:VCARD\n"
        "BEGIN:VCARD\n"
        "VERSION:3.0\n"
        "FN:Ori Bina\n"
        "N:Bina;Ori;;;\n"
        "TEL;TYPE=CELL:0520000000\n"
        "END:VCARD\n"
    ).encode("utf-8")


def _upload_valid_files(client: Any) -> Any:
    return client.post(
        "/upload",
        data={
            "workbook": (BytesIO(_make_workbook_bytes()), "guests.xlsx"),
            "contacts": [(BytesIO(_make_contacts_vcf()), "contacts.vcf")],
        },
        content_type="multipart/form-data",
        follow_redirects=True,
    )


def test_flask_upload_review_resolve_and_download_flow() -> None:
    app = create_app({"TESTING": True})

    client = app.test_client()
    upload_response = _upload_valid_files(client)

    assert upload_response.status_code == 200
    page = upload_response.get_data(as_text=True)
    assert 'dir="rtl"' in page
    assert "בחירת איש הקשר הנכון" in page
    assert "דורשים בדיקה" in page
    assert "הורדת קובץ אקסל" in page
    assert "review-" in page

    marker = 'name="choice_key" value="'
    choice_key = page.split(marker, 1)[1].split('"', 1)[0]

    resolve_response = client.post(
        "/resolve",
        data={"choice_key": choice_key, "candidate": "1"},
        follow_redirects=True,
    )

    assert resolve_response.status_code == 200
    resolved_page = resolve_response.get_data(as_text=True)
    assert "נפתר" in resolved_page
    assert '<div class="metric-value">1</div>' in resolved_page
    assert "הוכרע ידנית" in resolved_page

    download_response = client.get("/download")

    assert download_response.status_code == 200
    assert download_response.mimetype == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "guests-matched.xlsx" in download_response.headers["Content-Disposition"]

    workbook_file = BytesIO(download_response.data)
    workbook = load_workbook(workbook_file)
    guests = workbook["Guests"]
    assert guests["B2"].value == "0525665352"
    assert guests["B2"].number_format == "@"
    assert guests["C2"].value == "matched"
    assert guests["F2"].value == "manual-review"

    summary = workbook["Summary"]
    metrics = {
        summary[f"A{row_number}"].value: summary[f"B{row_number}"].value
        for row_number in range(2, summary.max_row + 1)
    }
    assert metrics["matched"] == 1
    assert metrics["review"] == 0
    assert metrics["unmatched"] == 0


def test_upload_failure_hides_internal_parser_errors() -> None:
    app = create_app({"TESTING": True})
    client = app.test_client()

    response = client.post(
        "/upload",
        data={
            "workbook": (BytesIO(b"not-an-xlsx"), "guests.xlsx"),
            "contacts": [(BytesIO(_make_contacts_vcf()), "contacts.vcf")],
        },
        content_type="multipart/form-data",
        follow_redirects=True,
    )

    assert response.status_code == 400
    page = response.get_data(as_text=True)
    assert "לא הצלחנו לעבד את הקבצים שהועלו." in page
    assert "File is not a zip file" not in page


def test_expired_session_is_evicted_on_access() -> None:
    app = create_app(
        {
            "TESTING": True,
            "SESSION_TTL_SECONDS": 1,
            "SESSION_CLEANUP_INTERVAL_SECONDS": 0,
        }
    )
    client = app.test_client()

    upload_response = _upload_valid_files(client)
    assert upload_response.status_code == 200

    sid = next(iter(SESSIONS))
    SESSIONS[sid].last_accessed_at -= 5

    response = client.get("/review", follow_redirects=False)

    assert response.status_code == 302
    assert response.headers["Location"].endswith("/")
    assert sid not in SESSIONS