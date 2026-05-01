from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import BinaryIO

from openpyxl import Workbook, load_workbook

from .models import GuestRow, MatchResult
from .text import normalize_name


NAME_COLUMNS = ["שם", "שם מלא", "שם מוזמן"]
OUTPUT_COLUMNS = ["טלפון", "סטטוס התאמה", "איש קשר תואם", "ציון התאמה", "סיבת התאמה"]


@dataclass(slots=True)
class SheetData:
    columns: list[str]
    rows: list[dict[str, object]]


def load_guest_workbook(
    source: str | Path | BytesIO | BinaryIO,
) -> tuple[dict[str, SheetData], list[GuestRow]]:
    spreadsheet = load_workbook(source, data_only=True)
    workbook: dict[str, SheetData] = {}
    guest_rows: list[GuestRow] = []
    for worksheet in spreadsheet.worksheets:
        raw_rows = list(worksheet.iter_rows(values_only=True))
        if not raw_rows:
            raise ValueError(f"Sheet '{worksheet.title}' is missing a name column (expected one of: {', '.join(NAME_COLUMNS)})")

        columns = [_stringify_header(value) for value in raw_rows[0]]
        
        name_col_found = None
        for col in NAME_COLUMNS:
            if col in columns:
                name_col_found = col
                break

        if not name_col_found:
            raise ValueError(f"Sheet '{worksheet.title}' is missing a name column (expected one of: {', '.join(NAME_COLUMNS)})")

        sheet_rows: list[dict[str, object]] = []
        for row_index, values in enumerate(raw_rows[1:]):
            row = {
                column: values[column_index] if column_index < len(values) else None
                for column_index, column in enumerate(columns)
            }
            sheet_rows.append(row)

            raw_name = _stringify_name(row.get(name_col_found, ""))
            if not raw_name:
                continue
            guest_rows.append(
                GuestRow(
                    sheet_name=worksheet.title,
                    row_index=row_index,
                    raw_name=raw_name,
                    normalized_name=normalize_name(raw_name),
                    values=dict(row),
                )
            )
        workbook[worksheet.title] = SheetData(columns=columns, rows=sheet_rows)
    return workbook, guest_rows


def write_output_workbook(workbook: dict[str, SheetData], results: list[MatchResult]) -> bytes:
    result_by_row = {(result.guest.sheet_name, result.guest.row_index): result for result in results}
    output = BytesIO()
    spreadsheet = Workbook()
    spreadsheet.remove(spreadsheet.active)

    for sheet_name, sheet_data in workbook.items():
        worksheet = spreadsheet.create_sheet(title=sheet_name)
        export_columns = list(sheet_data.columns)
        for column in OUTPUT_COLUMNS:
            if column not in export_columns:
                export_columns.append(column)

        worksheet.append(export_columns)
        output_indexes = {column: export_columns.index(column) + 1 for column in OUTPUT_COLUMNS}

        for row_index, row in enumerate(sheet_data.rows):
            result = result_by_row.get((sheet_name, row_index))
            export_row = dict(row)
            if result is not None:
                export_row["טלפון"] = result.phone_number
                export_row["סטטוס התאמה"] = result.status
                export_row["איש קשר תואם"] = (
                    result.matched_contact.full_name if result.matched_contact else ""
                )
                export_row["ציון התאמה"] = result.confidence
                export_row["סיבת התאמה"] = result.reason

            worksheet.append([export_row.get(column) for column in export_columns])
            _format_output_columns(worksheet, worksheet.max_row, output_indexes)

        _autosize_columns(worksheet, export_columns)

    _build_summary_sheet(spreadsheet, results)
    spreadsheet.save(output)
    output.seek(0)
    return output.read()


def _build_summary_sheet(spreadsheet: Workbook, results: list[MatchResult]) -> None:
    matched = sum(1 for result in results if result.status == "matched")
    review = sum(1 for result in results if result.status == "review")
    unmatched = sum(1 for result in results if result.status == "unmatched")
    worksheet = spreadsheet.create_sheet(title="Summary")
    worksheet.append(["metric", "count"])
    worksheet.append(["matched", matched])
    worksheet.append(["review", review])
    worksheet.append(["unmatched", unmatched])
    worksheet.append(["total", len(results)])


def _stringify_name(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value != value:
        return ""
    return str(value).strip()


def _stringify_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _format_output_columns(
    worksheet: Any,
    row_number: int,
    output_indexes: dict[str, int],
) -> None:
    for column_name in ("טלפון", "סטטוס התאמה", "איש קשר תואם", "סיבת התאמה"):
        column_number = output_indexes[column_name]
        cell = worksheet.cell(row=row_number, column=column_number)
        if cell.value is None:
            continue
        cell.value = str(cell.value)
        cell.number_format = "@"


def _autosize_columns(worksheet: Any, columns: list[str]) -> None:
    for column_number, column_name in enumerate(columns, start=1):
        max_length = len(str(column_name))
        for row in worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row,
            min_col=column_number,
            max_col=column_number,
        ):
            value = row[0].value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        worksheet.column_dimensions[worksheet.cell(row=1, column=column_number).column_letter].width = min(max_length + 2, 40)